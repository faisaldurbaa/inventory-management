import os
import tkinter as tk
from datetime import date
from datetime import datetime
from tkinter import *
from tkinter import ttk
import customtkinter as ctk
import openpyxl
from PIL import Image
from openpyxl import *
import sys

# data file aspects
base_path = getattr(sys, '_MEIPASS', os.path.abspath(os.path.dirname(__file__)))
file = "database.xlsx"
sheet1 = "İşlemler"
sheet2 = "Stok"
sheet3 = "Geçmiş"
trs_columns = ["İşlem No", "Ürün No", "Ürün", "Marka", "Birim", "Miktar", "İşlem Tarihi", "STT"]
stk_columns = ["Ürün No", "Ürün", "Marka", "Birim", "Miktar", "Son İşlem Tarihi", "STT"]
hst_columns = ["İşlem No", "Ürün No", "Ürün", "Marka", "Birim", "Miktar", "İşlem Tarihi", "STT", "Durum"]


def file_control(filename, sheet_1, header_1, sheet_2, header_2, sheet_3, header_3, stock=None, history=None):
    files = os.listdir(os.getcwd())
    if filename in files:
        return
    else:
        # If the file doesn't exist, create a new workbook and format it
        wb = Workbook()
        sheet_one = wb.create_sheet(title=sheet_1)
        sheet_two = wb.create_sheet(title=sheet_2)
        sheet_three = wb.create_sheet(title=sheet_3)
        wb.save(filename)

        sheet_one.append(header_1)
        sheet_two.append(header_2)
        if type(stock) is type(None):
            pass
        else:
            for i in stock[1:]:
                sheet_two.append(i)

        sheet_three.append(header_3)
        if type(stock) is type(None):
            pass
        else:
            for n in history[1:]:
                sheet_three.append(n)

        wb.remove(wb['Sheet'])

        wb.save(filename)
        wb.close()


def data(filename, sheet):  # return a nested-list version of the database
    wb = load_workbook(filename, read_only=True)
    data_sheet = wb[sheet]
    data_list = [list(row) for row in data_sheet.iter_rows(values_only=True)]
    wb.close()
    return data_list


def next_operation_code(file_data):
    if file_data[1:]:
        largest_code = max(row[0] for row in file_data[1:])  # Find the largest operation code
        new_operation_code = int(largest_code) + 1
        return new_operation_code
    else:
        return int(1)  # Start with operation code 1 if no existing data


def next_product_code(file_data, name):
    if file_data[1:]:
        product_codes = [row[0] for row in file_data[1:]]
        names = [row[1] for row in file_data[1:]]
        if name in names:
            code = product_codes[names.index(name)]
            return code

        else:
            code_letters = [code[0:2] for code in product_codes]
            max_code_letters = max(code_letters)
            code_numbers = [code[2:4] for code in product_codes]
            max_code_number = max(code_numbers)
            number, letter1, letter2 = max_code_number, max_code_letters[0], max_code_letters[1]

            if int(max_code_number) == 99:
                number = "00"
                if max_code_letters[1] == 'Z':
                    letter2 = "A"
                    if max_code_letters[0] == 'Z':
                        raise ValueError("Item code sequence limit reached.")
                    else:
                        letter1 = chr(ord(max_code_letters[0] + 1))
                else:
                    letter2 = chr(ord(max_code_letters[1] + 1))
            else:
                if number[0] == '0' and number[1] == '9':
                    number = 10
                elif number[0] == '0':
                    number = f"0{int(number) + 1}"
                else:
                    number = int(number) + 1

            new_code = letter1 + letter2 + str(number)

            return new_code

    else:
        return "AA01"


def write_new_row(row_data, file_name, sheet):
    wb = openpyxl.load_workbook(file_name)
    w_sheet = wb[sheet]
    w_sheet.append(row_data)
    wb.save(file_name)
    wb.close()


def overwrite(new_data, index, file_name, sheet):
    wb = openpyxl.load_workbook(file_name)
    ac_sheet = wb[sheet]
    for col_num, value in enumerate(new_data, start=1):
        ac_sheet.cell(row=index, column=col_num, value=value)
    wb.save(file_name)
    wb.close()


def totals_from_zero(file_name, sheet_1, sheet_2):
    file_data = data(file_name, sheet_1)

    unique = [list(element) for element in list({tuple([row[1], row[3], row[4], row[7]]) for row in file_data[1:]})]

    for uni in unique:
        for row in file_data[1:]:
            if row[1] == uni[0] and row[3] == uni[1] and row[4] == uni[2] and row[7] == uni[3]:
                uni.append(row)

    total_values = []
    for part in unique:
        t_quantity = 0
        last_opr = []
        for ls in part[4:]:
            t_quantity += float(ls[5])
            last_opr.append(ls[6])
        name = part[4][2]
        part.insert(1, name)
        part.insert(4, t_quantity)
        part.insert(5, max(last_opr))
        total_values.append(part[:7])

    wb = openpyxl.load_workbook(file_name)
    sheet = wb[sheet_2]
    for row in range(2, sheet.max_row + 1):
        sheet.delete_rows(2)
    wb.save(file_name)
    wb.close()

    for i in total_values:
        write_new_row(i, file_name, sheet_2)


def new_to_totals(new_data, file_name, sheet):
    file_data = data(file_name, sheet)
    done = False
    for row in file_data[1:]:
        if new_data[1] == row[0] and new_data[3] == row[2] and new_data[4] == row[3] and new_data[7] == row[6]:
            n_quantity = float(row[4]) + float(new_data[5])
            if new_data[6] > row[5]:
                last_opr = new_data[6]
            else:
                last_opr = row[5]
            new_row_data = [row[0], row[1], row[2], row[3], n_quantity, last_opr, row[6]]
            overwrite(new_row_data, file_data.index(row) + 1, file_name, sheet)
            done = True

    if not done:
        write_new_row(new_data[1:], file_name, sheet)


def file_output(source, ssh1, ssh2, ssh3):
    sh1 = data(source, ssh1)
    sh2 = data(source, ssh2)
    sh3 = data(source, ssh3)

    wb = Workbook()
    sheet_one = wb.create_sheet(title=ssh1)
    for i in sh1:
        sheet_one.append(i)
    sheet_two = wb.create_sheet(title=ssh2)
    for n in sh2:
        sheet_two.append(n)
    sheet_three = wb.create_sheet(title=ssh3)
    for m in sh3:
        sheet_three.append(m)
    wb.remove(wb['Sheet'])
    wb.save(rf"{datetime.now().strftime('%Y%m%d%H%M%S')}_raporu.xlsx")
    wb.close()


def delete_row(file_name, sheet, row):
    wb = openpyxl.load_workbook(file_name)
    ac_sheet = wb[sheet]
    ac_sheet.delete_rows(row)
    wb.save(file_name)
    wb.close()


def create_data_entry_frame(master):
    frame = ctk.CTkFrame(master, fg_color='#2b2b2b')

    frame.grid_columnconfigure(0, weight=6)
    frame.grid_columnconfigure(1, weight=4)
    frame.grid_rowconfigure(0, weight=4)

    # Data entry frame
    entry_frame = ctk.CTkFrame(frame, fg_color='#333333')
    entry_frame.grid(row=0, column=0, sticky="ns", pady=frame.winfo_reqheight()*1/10)

    space_x = int(entry_frame.winfo_reqwidth()*1/25)
    space_y = int(entry_frame.winfo_reqheight()*1/15)

    def update_paddings(event):
        window_width = event.width

        # Adjust the padding between frames based on the window width
        padding_between_frames = int(window_width * 0.02)

        # Apply the new padding between frames
        entry_frame.grid(padx=padding_between_frames)
        search_frame.grid(padx=padding_between_frames)


    # Add the data entry boxes to the frame
    entry_label = ctk.CTkLabel(entry_frame, text="Veri Girişi", font=("Arial", 22, "bold"),
                               fg_color="gray30", corner_radius=6, text_color='#ffffff')
    entry_label.configure(height=50)
    entry_label.grid(row=0, column=2, columnspan=3, sticky="nsew", pady=(space_y*0.5, space_y), padx=space_x)

    name_label = ctk.CTkLabel(entry_frame, text="Ürün", font=("Arial", 18, "bold"))
    name_label.grid(row=1, column=0, sticky="e", padx=(2*space_x,space_x), pady=space_y)
    name_entry = ctk.CTkEntry(entry_frame, font=("Arial", 16, "normal"))
    name_entry.configure(height=30)
    name_entry.grid(row=1, column=1, pady=space_y, padx=(0, 2*space_x))

    brand_label = ctk.CTkLabel(entry_frame, text="Marka", font=("Arial", 18, "bold"))
    brand_label.grid(row=1, column=2, sticky="e", padx=(0,space_x), pady=space_y)
    brand_entry = ctk.CTkEntry(entry_frame, font=("Arial", 16, "normal"))
    brand_entry.configure(height=30)
    brand_entry.grid(row=1, column=3, pady=space_y, padx=(0, 2*space_x))

    expdate_label = ctk.CTkLabel(entry_frame, text="STT", font=("Arial", 18, "bold"))
    expdate_label.grid(row=1, column=4, sticky="e", padx=(0, int(space_x/4)), pady=space_y)
    day_entry = ctk.CTkEntry(entry_frame, font=("Arial", 16, "normal"))
    day_entry.configure(height=30, width=30)
    day_entry.grid(row=1, column=5, pady=space_y)
    month_entry = ctk.CTkEntry(entry_frame, font=("Arial", 16, "normal"))
    month_entry.configure(height=30, width=30)
    month_entry.grid(row=1, column=6, padx=(0,int(space_x/4)), pady=space_y)
    year_entry = ctk.CTkEntry(entry_frame, font=("Arial", 16, "normal"))
    year_entry.configure(height=30, width=55)
    year_entry.grid(row=1, column=7, pady=space_y, padx=(0,2*space_x))

    unit_label = ctk.CTkLabel(entry_frame, text="Birim", font=("Arial", 18, "bold"))
    unit_label.grid(row=2, column=0, sticky="e", padx=(2*space_x, space_x), pady=space_y)
    unit_dropdown = ctk.CTkComboBox(entry_frame, values=["", "kg", "L", "Şişe", "Kutu"], font=("Arial", 16, "normal"))
    unit_dropdown.configure(height=30)
    unit_dropdown.grid(row=2, column=1, pady=space_y, padx=(0, 2*space_x))

    quantity_label = ctk.CTkLabel(entry_frame, text="Miktar", font=("Arial", 18, "bold"))
    quantity_label.grid(row=2, column=2, sticky="e", padx=(0,space_x), pady=space_y)
    quantity_entry = ctk.CTkEntry(entry_frame, font=("Arial", 16, "normal"))
    quantity_entry.configure(height=30)
    quantity_entry.grid(row=2, column=3, pady=space_y, padx=(0, 2*space_x))

    # Search (code) frame
    search_frame = ctk.CTkFrame(frame, fg_color='#333333')
    search_frame.grid(column=1, row=0, sticky="wns", padx=0, pady=frame.winfo_reqheight()*1/10)

    frame.grid_columnconfigure(0, weight=2)
    frame.grid_columnconfigure(1, weight=2)
    frame.grid_columnconfigure(2, weight=0)

    # labels and entry boxes
    entry_label = ctk.CTkLabel(search_frame, text="Arama Kodları", font=("Arial", 22, "bold"),
                               fg_color="gray30", corner_radius=6, text_color='#ffffff')
    entry_label.configure(height=50)
    entry_label.grid(row=0, column=0, columnspan=4, pady=(space_y*0.5, space_y), padx=space_x/2,  ipadx=5)

    procode_label = ctk.CTkLabel(search_frame, text="Ürün Kodu", font=("Arial", 18, "bold"))
    procode_label.grid(row=1, column=0, padx=int(search_frame.winfo_reqwidth()/25), sticky="w", pady= int(search_frame.winfo_reqheight()/15))
    procode_entry = ctk.CTkEntry(search_frame, font=("Arial", 16, "normal"))
    procode_entry.configure(height=30, width=100)
    procode_entry.grid(row=1, column=1, padx=(0, 20), pady= int(search_frame.winfo_reqheight()/15))

    oprcode_label = ctk.CTkLabel(search_frame, text="İşlem Kodu", font=("Arial", 18, "bold"))
    oprcode_label.grid(row=2, column=0, padx=int(search_frame.winfo_reqwidth()/25), sticky="w", pady= int(search_frame.winfo_reqheight()/15))
    oprcode_entry = ctk.CTkEntry(search_frame, font=("Arial", 16, "normal"))
    oprcode_entry.configure(height=30, width=100)
    oprcode_entry.grid(row=2, column=1, padx=(0, 20), pady=int(search_frame.winfo_reqheight()/15))

    def empty():
        name_entry.delete(0, 'end')
        brand_entry.delete(0, 'end')
        day_entry.delete(0, 'end')
        month_entry.delete(0, 'end')
        year_entry.delete(0, 'end')
        unit_dropdown.set('')
        quantity_entry.delete(0, 'end')
        procode_entry.delete(0, 'end')
        oprcode_entry.delete(0, 'end')

    frame.empty = empty

    empty_button = ctk.CTkButton(entry_frame, text="Hepsini Sil", font=("Arial", 16, "bold"), command=empty)
    empty_button.configure(height=entry_frame.winfo_reqheight()*0.15)
    empty_button.grid(row=2, column=5, columnspan=3, pady=20, padx=(0,space_x*2), sticky="we")

    spacing = int(frame.winfo_reqwidth() - (search_frame.winfo_reqwidth() + entry_frame.winfo_reqwidth()) / 3)
    entry_frame.grid_configure(padx=spacing)
    search_frame.grid_configure(padx=(0, spacing))


    # Assign widgets to the frame attributes
    frame.entry_label = entry_label
    frame.name_label = name_label
    frame.name_entry = name_entry
    frame.brand_label = brand_label
    frame.brand_entry = brand_entry
    frame.expdate_label = expdate_label
    frame.day_entry = day_entry
    frame.month_entry = month_entry
    frame.year_entry = year_entry
    frame.unit_label = unit_label
    frame.unit_dropdown = unit_dropdown
    frame.quantity_label = quantity_label
    frame.quantity_entry = quantity_entry
    frame.procode_label = procode_label
    frame.procode_entry = procode_entry
    frame.oprcode_label = oprcode_label
    frame.oprcode_entry = oprcode_entry

    frame.bind("<Configure>", update_paddings)

    return frame


def create_button_frame(master):
    frame = ctk.CTkFrame(master, fg_color='#2b2b2b')

    frame.grid_columnconfigure(0, weight=1)
    frame.grid_columnconfigure(1, weight=10)
    frame.grid_columnconfigure(2, weight=1)
    frame.grid_rowconfigure(0, weight=2)
    frame.grid_rowconfigure(1, weight=8)
    frame.grid_rowconfigure(2, weight=3)

    # Construct the full path to the photo
    photo_path = os.path.join(base_path, 'VDFL_logo.png')

    logo_image = ctk.CTkImage(dark_image=Image.open(photo_path), size=(int(frame.winfo_reqwidth()*0.7), int(frame.winfo_reqwidth()*0.7)))
    logo_label = ctk.CTkLabel(master=frame, image=logo_image, text="")
    logo_label.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")

    # data buttons frame
    data_buttons_frame = ctk.CTkFrame(frame, fg_color='#333333')
    data_buttons_frame.grid(row=1, column=1, padx=10, pady=0, sticky="nswe")
    data_buttons_frame.configure(corner_radius=0)

    data_operations_label = ctk.CTkLabel(data_buttons_frame, text="Veri İşlemleri", fg_color="gray30", corner_radius=6,
                                         text_color='#ffffff')
    data_operations_label.grid(row=0, column=0, sticky="nsew", pady=5, padx=5)
    data_operations_label.configure(font=("Arial", 22, "bold"))

    # noinspection PyUnboundLocalVariable
    def select_opr():
        # empty the entries
        if radio_var.get() != 3 and master.dt_frame.db1.focused() != "":
            master.de_frame.name_entry.delete(0, 'end')
            master.de_frame.brand_entry.delete(0, 'end')
            master.de_frame.day_entry.delete(0, 'end')
            master.de_frame.month_entry.delete(0, 'end')
            master.de_frame.year_entry.delete(0, 'end')
            master.de_frame.unit_dropdown.set('')
            master.de_frame.quantity_entry.delete(0, 'end')
            master.de_frame.procode_entry.delete(0, 'end')
            master.de_frame.oprcode_entry.delete(0, 'end')

            # get the selected row
            selected = master.dt_frame.db1.focused()
            values = master.dt_frame.db1.val(selected)

            # values to entry boxes
            if radio_var.get() == 1:
                master.de_frame.name_entry.insert(0, values[2])
                master.de_frame.brand_entry.insert(0, values[3])
                master.de_frame.day_entry.insert(0, values[7][0:2])
                master.de_frame.month_entry.insert(0, values[7][3:5])
                master.de_frame.year_entry.insert(0, values[7][6:10])
                master.de_frame.unit_dropdown.set(values[4])
                master.de_frame.quantity_entry.insert(0, values[5])
                master.de_frame.procode_entry.insert(0, values[1])
                master.de_frame.oprcode_entry.insert(0, values[0])
            elif radio_var.get() == 2:
                master.de_frame.name_entry.insert(0, values[1])
                master.de_frame.brand_entry.insert(0, values[2])
                master.de_frame.day_entry.insert(0, values[6][0:2])
                master.de_frame.month_entry.insert(0, values[6][3:5])
                master.de_frame.year_entry.insert(0, values[6][6:10])
                master.de_frame.unit_dropdown.set(values[3])
                master.de_frame.quantity_entry.insert(0, values[4])
                master.de_frame.procode_entry.insert(0, values[0])

    def errorbox(text):
        error_box = ctk.CTkToplevel(master)
        error_box.wm_transient(master=master)
        width = len(text) * 8.4
        error_box.geometry(f"{width}x100")
        error_box.title("Error")
        label = ctk.CTkLabel(error_box, text=text, font=('Arial', 18))
        label.grid(column=0, row=0, pady=5, padx=5)
        close = ctk.CTkButton(error_box, text='Tamam', command=error_box.destroy, font=('Arial', 18))
        close.grid(column=0, row=1, pady=5, padx=5)
        error_box.lift()

    def data_input():
        try:
            name = master.de_frame.name_entry.get().strip().capitalize()
            brand = master.de_frame.brand_entry.get().strip().capitalize()
            lst = list(map(int, [entry.get() for entry in [master.de_frame.year_entry, master.de_frame.month_entry,
                                                           master.de_frame.day_entry]]))
            expiry = date(lst[0], lst[1], lst[2]).strftime("%d/%m/%Y")
            unit = master.de_frame.unit_dropdown.get().strip() if (master.de_frame.unit_dropdown.get().strip()
                                                                   in ["kg", "L", "Şişe", "Kutu"]) else None
            quantity = float(master.de_frame.quantity_entry.get())
        except ValueError:
            errorbox("Lütfen Tekrar Deneyiniz!")
        else:
            if (None or "") in [name, brand, unit]:
                errorbox("Lütfen Tekrar Deneyiniz!")
            else:
                f_data = data(file, sheet1)
                opr_code = next_operation_code(f_data)
                total_data = data(file, sheet2)
                pro_code = next_product_code(total_data, name)
                current_date = date.today().strftime("%d/%m/%Y")
                new_data = [opr_code, pro_code, name, brand, unit, quantity, current_date, expiry]
                write_new_row(new_data, file, sheet1)
                new_to_totals(new_data, file, sheet2)
                master.de_frame.empty()
                update_data_tables_frame()

    def report():
        file_output(file, sheet1, sheet2, sheet3)

    def output():
        try:
            name = master.de_frame.name_entry.get().strip().capitalize()
            brand = master.de_frame.brand_entry.get().strip().capitalize()
            lst = list(map(int, [entry.get() for entry in [master.de_frame.year_entry, master.de_frame.month_entry,
                                                           master.de_frame.day_entry]]))
            expiry = date(lst[0], lst[1], lst[2]).strftime("%d/%m/%Y")
            unit = master.de_frame.unit_dropdown.get().strip() if (master.de_frame.unit_dropdown.get().strip()
                                                                   in ["kg", "L", "Şişe", "Kutu"]) else None
            quantity = float(master.de_frame.quantity_entry.get().strip())
        except ValueError:
            errorbox("Lütfen Tekrar Deneyiniz!")
        else:
            if (None or "") in [name, brand, unit]:
                errorbox("Lütfen Tekrar Deneyiniz!")
            else:
                stock_data = data(file, sheet2)
                trs_data = data(file, sheet1)
                found = False
                for row in stock_data[1:]:
                    if row[1] == name and row[2] == brand and row[3] == unit and row[6] == expiry:
                        # expiry is the problem
                        found = True
                        available = float(row[4])
                        if available < quantity:
                            return errorbox(f"Bu Üründen {available} Adet Bulunmaktadır.")
                        else:
                            n_qt = available - quantity
                            lst = []
                            for i in trs_data[1:]:
                                if i[1] == row[0] and i[3] == row[2] and i[4] == row[3] and i[7] == row[6]:
                                    lst.append(i[6])

                            last_opr = max(lst) if len(lst) > 0 else "00/00/0000"
                            new_row_data = [row[0], row[1], row[2], row[3], n_qt, last_opr, row[6]]
                            overwrite(new_row_data, stock_data.index(row) + 1, file, sheet2)
                            opr = next_operation_code(trs_data)
                            pro = row[0]
                            current_date = date.today().strftime("%d/%m/%Y")
                            new = [opr, pro, name, brand, unit, -quantity, current_date, expiry]
                            write_new_row(new, file, sheet1)
                            master.de_frame.empty()
                            update_data_tables_frame()

                if not found:
                    return errorbox("Ürün Bulunamadı!")

    def edit():
        try:
            name = master.de_frame.name_entry.get().strip().capitalize()
            brand = master.de_frame.brand_entry.get().strip().capitalize()
            lst = list(
                map(int, [entry.get().strip() for entry in [master.de_frame.year_entry, master.de_frame.month_entry,
                                                            master.de_frame.day_entry]]))
            expiry = date(lst[0], lst[1], lst[2]).strftime("%d/%m/%Y")
            unit = master.de_frame.unit_dropdown.get().strip() if (master.de_frame.unit_dropdown.get().strip()
                                                                   in ["kg", "L", "Şişe", "Kutu"]) else None
            quantity = float(master.de_frame.quantity_entry.get().strip())
            opr_code = int(master.de_frame.oprcode_entry.get().strip())
        except ValueError:
            return errorbox("Tekrar Deneyiniz")
        else:
            if (None or "" or 0) in [name, brand, unit, quantity]:
                return errorbox("Tekrar Deneyiniz")

            trs = data(file, sheet1)
            stk = data(file, sheet2)
            done = False
            for line in trs[1:]:
                if int(line[0]) == opr_code:  # add else for "not found"
                    #  check if the updated product exists or if the amount is sufficient when making output.
                    if quantity < 0:
                        found = False
                        for n in stk[1:]:
                            if n[1] == name and n[2] == brand and n[3] == unit and n[6] == expiry:
                                found = True
                                available = float(n[4]) - float(
                                    line[5])  # extract the formerly entered quantity from the shown one
                                if available < -quantity:
                                    return errorbox(
                                        f"Bu Üründen {available} ({float(n[4])} + {- float(line[5])})"
                                        f" Adet Bulunmaktadır.")
                        if not found:
                            return errorbox("Ürün Bulunamadı!")

                    done = True
                    # undo the operation, and update last operation date and quantity for the product:
                    for row in stk[1:]:
                        if row[0] == line[1]:  # don't add else
                            nqt = float(row[4]) - float(line[5])
                            lst = []
                            for i in trs[1:]:
                                if i[1] == row[0] and i[3] == row[2] and i[4] == row[3] and i[7] == row[6]:
                                    lst.append(i[6])

                            last_opr = max(lst) if len(lst) > 0 else "00/00/0000"
                            new_row = [row[0], row[1], row[2], row[3], nqt, last_opr, row[6]]
                            overwrite(new_row, stk.index(row) + 1, file, sheet2)

                    # write the first version of the operation to history sheet:
                    line.append("Düzenlendi")
                    write_new_row(line, file, sheet3)
                    line.remove("Düzenlendi")

                    # take the data entry again and edit the operation.
                    trs = data(file, sheet1)
                    stk = data(file, sheet2)
                    if quantity < 0:
                        for n in stk[1:]:
                            if n[1] == name and n[2] == brand and n[3] == unit and n[6] == expiry:
                                available = float(n[4])
                                n_qt = available + quantity
                                pro = n[0]
                                current_date = date.today().strftime("%d/%m/%Y")
                                new_row_data = [n[0], n[1], n[2], n[3], n_qt, current_date, n[6]]
                                overwrite(new_row_data, stk.index(n) + 1, file, sheet2)
                                new = [opr_code, pro, name, brand, unit, quantity, current_date, expiry]
                                overwrite(new, trs.index(line) + 1, file, sheet1)
                                master.de_frame.empty()
                                update_data_tables_frame()
                    else:
                        total_data = data(file, sheet2)
                        pro_code = next_product_code(total_data, name)
                        current_date = date.today().strftime("%d/%m/%Y")
                        new_data = [opr_code, pro_code, name, brand, unit, quantity, current_date, expiry]
                        overwrite(new_data, trs.index(line) + 1, file, sheet1)
                        new_to_totals(new_data, file, sheet2)
                        master.de_frame.empty()
                        update_data_tables_frame()

            if not done:
                return errorbox("İşlem Bulunamadı")

    def delete():
        try:
            opr = int(master.de_frame.oprcode_entry.get().strip())
        except ValueError:
            return errorbox("İşlem Kodunu Doğru Giriniz!")

        trs = data(file, sheet1)
        stk = data(file, sheet2)
        found = False
        for ln in trs[1:]:
            if int(ln[0]) == opr:
                found = True
                for row in stk[1:]:
                    if row[0] == ln[1]:  # don't add else
                        nqt = float(row[4]) - float(ln[5])
                        lst = []
                        for i in trs[1:]:
                            if i[1] == row[0] and i[3] == row[2] and i[4] == row[3] and i[7] == row[6]:
                                lst.append(i[6])

                        last_opr = max(lst) if len(lst) > 0 else "00/00/0000"
                        new_row = [row[0], row[1], row[2], row[3], nqt, last_opr, row[6]]
                        overwrite(new_row, stk.index(row) + 1, file, sheet2)

                ln.append("Silindi")
                write_new_row(ln, file, sheet3)
                ln.remove("Silindi")

                delete_row(file, sheet1, trs.index(ln) + 1)

                master.de_frame.empty()
                update_data_tables_frame()

        if not found:
            return errorbox('İşlem Bulunamadı!')

    # buttons
    button1 = ctk.CTkButton(data_buttons_frame, text="İşlemi Seç", font=("Arial", 18, "bold"), command=select_opr)
    button1.grid(row=1, column=0, padx=10, pady=(10, 0), sticky="nsew")

    button2 = ctk.CTkButton(data_buttons_frame, text="Ürün Girişi", font=("Arial", 18, "bold"), command=data_input)
    button2.grid(row=2, column=0, padx=10, pady=10, sticky="nsew")

    button3 = ctk.CTkButton(data_buttons_frame, text="Ürün Çıkışı", font=("Arial", 18, "bold"), command=output)
    button3.grid(row=3, column=0, padx=10, pady=(0, 10), sticky="nsew")

    button4 = ctk.CTkButton(data_buttons_frame, text="İşlemi Düzenle", font=("Arial", 18, "bold"), command=edit)
    button4.grid(row=4, column=0, padx=10, pady=(0, 10), sticky="nsew")

    button5 = ctk.CTkButton(data_buttons_frame, text="İşlemi Sil", font=("Arial", 18, "bold"), command=delete)
    button5.grid(row=5, column=0, padx=10, pady=(0, 10), sticky="nsew")

    button6 = ctk.CTkButton(data_buttons_frame, text="Rapor Oluştur", font=("Arial", 18, "bold"), fg_color="#009900",
                            command=report)
    button6.grid(row=6, column=0, padx=10, pady=(0, 10), sticky="nsew")

    data_buttons_frame.grid_columnconfigure(0, weight=1)
    data_buttons_frame.grid_rowconfigure(0, weight=2)
    data_buttons_frame.grid_rowconfigure(1, weight=2)
    data_buttons_frame.grid_rowconfigure(2, weight=2)
    data_buttons_frame.grid_rowconfigure(3, weight=2)
    data_buttons_frame.grid_rowconfigure(4, weight=2)
    data_buttons_frame.grid_rowconfigure(5, weight=2)
    data_buttons_frame.grid_rowconfigure(6, weight=2)

    # database buttons frame
    database_buttons_frame = ctk.CTkFrame(frame, fg_color='#333333')
    database_buttons_frame.grid(row=2, column=1, padx=10, pady=10, sticky="nswe")
    database_buttons_frame.configure(corner_radius=0)

    databases_label = ctk.CTkLabel(database_buttons_frame, text="Veri Tabanları", fg_color="gray30", corner_radius=6, text_color='#ffffff')
    databases_label.grid(row=0, column=0, sticky="nsew", pady=5, padx=5)
    databases_label.configure(font=("Arial", 22, "bold"))

    radio_var = tk.IntVar(value=1)

    def update_data_tables_frame():
        selected_db = radio_var.get()
        if selected_db == 1:
            master.dt_frame.db1()
        elif selected_db == 2:
            master.dt_frame.db2()
        elif selected_db == 3:
            master.dt_frame.db3()

    update_data_tables_frame()

    button_1 = ctk.CTkRadioButton(database_buttons_frame, text='İşlemler', font=("Arial", 19, "bold"),
                                  variable=radio_var, value=1, command=update_data_tables_frame)
    button_1.grid(row=1, column=0, padx=10, sticky="nw")

    button_2 = ctk.CTkRadioButton(database_buttons_frame, text='Stok', font=("Arial", 19, "bold"),
                                  variable=radio_var, value=2, command=update_data_tables_frame)
    button_2.grid(row=2, column=0, padx=10, sticky="nw")

    button_3 = ctk.CTkRadioButton(database_buttons_frame, text='İşlem Geçmişi', font=("Arial", 19, "bold"),
                                  variable=radio_var, value=3, command=update_data_tables_frame)
    button_3.grid(row=3, column=0, padx=10, sticky="nw")

    database_buttons_frame.grid_columnconfigure(0, weight=1)
    database_buttons_frame.grid_rowconfigure(0, weight=2)
    database_buttons_frame.grid_rowconfigure(1, weight=2)
    database_buttons_frame.grid_rowconfigure(2, weight=2)

    return frame


def create_data_tables_frame(master):
    frame = ctk.CTkFrame(master)

    # Keep the data, columns, and treeview as attributes
    dt_frame_data = None
    dt_frame_columns = None
    dt_frame_treeview = None
    dt_frame_scrollbar = None

    # noinspection PyUnresolvedReferences
    def db1():
        nonlocal dt_frame_treeview, dt_frame_data, dt_frame_columns, dt_frame_scrollbar
        if dt_frame_treeview:
            # noinspection PyUnresolvedReferences
            dt_frame_treeview.destroy()
        if dt_frame_scrollbar:
            dt_frame_scrollbar.destroy()

        # get the data
        dt_frame_data = data(file, sheet1)
        dt_frame_columns = dt_frame_data[0]

        # Create a style for the Treeview
        style = ttk.Style()
        style.theme_use("default")  # You can experiment with different themes (classic, clam, alt, default)

        style.configure("Treeview", background="#D3D3D3", fieldbackground="#D3D3D3",
                        foreground="black", rowheight=25, font=('Arial', 14))

        style.configure('Treeview.Heading', font=('Arial', 14, "bold"))

        style.map('Treeview', background=[('selected', "#347083")])

        # scrollbar
        dt_frame_scrollbar = ctk.CTkScrollbar(frame)
        dt_frame_scrollbar.pack(side='right', fill='y')
        # Create the Treeview with the configured style
        dt_frame_treeview = ttk.Treeview(frame, columns=dt_frame_columns, selectmode='extended',
                                         show="headings", style="Treeview",
                                         yscrollcommand=dt_frame_scrollbar.set)

        dt_frame_treeview.tag_configure('oddrow', background='#ffffff')
        dt_frame_treeview.tag_configure('evenrow', background='lightblue')

        dt_frame_scrollbar.configure(command=dt_frame_treeview.yview)

        # Add column headings and data
        for col in dt_frame_columns:
            dt_frame_treeview.heading(col, text=col)
            dt_frame_treeview.column(col, width=int(master.winfo_width()*4/6*1/8), anchor='center')

        for row in dt_frame_data[1:]:
            row[7] = str(row[7])[:10]

        count = 0
        for row in dt_frame_data[1:]:
            if count % 2:
                dt_frame_treeview.insert("", tk.END, values=row, tags="evenrow")
            else:
                dt_frame_treeview.insert("", tk.END, values=row, tags="oddrow")
            count += 1

        def focused():
            focus = dt_frame_treeview.focus()
            return focus

        def val(selected):
            values = dt_frame_treeview.item(selected, 'values')
            return values

        db1.focused = focused
        db1.val = val

        dt_frame_treeview.pack(fill=tk.BOTH, expand=True)

    # noinspection PyUnresolvedReferences
    def db2():
        nonlocal dt_frame_treeview, dt_frame_data, \
            dt_frame_columns, dt_frame_scrollbar
        if dt_frame_treeview:
            dt_frame_treeview.destroy()
        if dt_frame_scrollbar:
            dt_frame_scrollbar.destroy()

        # get the data
        dt_frame_data = data(file, sheet2)
        dt_frame_columns = dt_frame_data[0]

        # Create a style for the Treeview
        style = ttk.Style()
        style.theme_use("default")  # You can experiment with different themes (classic, clam, alt, default)

        style.configure("Treeview", background="#D3D3D3", fieldbackground="#D3D3D3",
                        foreground="black", rowheight=25, font=('Arial', 14))

        style.configure('Treeview.Heading', font=('Arial', 14, "bold"))

        style.map('Treeview', background=[('selected', "#347083")])

        # scrollbar
        dt_frame_scrollbar = ctk.CTkScrollbar(frame)
        dt_frame_scrollbar.pack(side='right', fill='y')
        # Create the Treeview with the configured style
        dt_frame_treeview = ttk.Treeview(frame, columns=dt_frame_columns, selectmode='extended',
                                         show="headings", style="Treeview",
                                         yscrollcommand=dt_frame_scrollbar.set)

        dt_frame_treeview.tag_configure('oddrow', background='#ffffff')
        dt_frame_treeview.tag_configure('evenrow', background='lightblue')

        dt_frame_scrollbar.configure(command=dt_frame_treeview.yview)

        # Add column headings and data
        for col in dt_frame_columns:
            dt_frame_treeview.heading(col, text=col)
            dt_frame_treeview.column(col, width=int(master.winfo_width()*4/6*1/7), anchor='center')

        for row in dt_frame_data[1:]:
            row[6] = str(row[6])[:10]
            row[5] = str(row[5])[:10]

        count = 0
        for row in dt_frame_data[1:]:
            if count % 2:
                dt_frame_treeview.insert("", tk.END, values=row, tags="evenrow")
            else:
                dt_frame_treeview.insert("", tk.END, values=row, tags="oddrow")
            count += 1

        def focused():
            focus = dt_frame_treeview.focus()
            return focus

        def val(selected):
            values = dt_frame_treeview.item(selected, 'values')
            return values

        db1.focused = focused
        db1.val = val

        dt_frame_treeview.pack(fill=tk.BOTH, expand=True)

    # noinspection PyUnresolvedReferences
    def db3():
        nonlocal dt_frame_treeview, dt_frame_data, \
            dt_frame_columns, dt_frame_scrollbar
        if dt_frame_treeview:
            dt_frame_treeview.destroy()
        if dt_frame_scrollbar:
            dt_frame_scrollbar.destroy()

        # get the data
        dt_frame_data = data(file, sheet3)
        dt_frame_columns = dt_frame_data[0]

        # Create a style for the Treeview
        style = ttk.Style()
        style.theme_use("default")  # You can experiment with different themes (classic, clam, alt, default)

        style.configure("Treeview", background="#D3D3D3", fieldbackground="#D3D3D3",
                        foreground="black", rowheight=25, font=('Arial', 14))

        style.configure('Treeview.Heading', font=('Arial', 14, "bold"))

        style.map('Treeview', background=[('selected', "#347083")])

        # scrollbar
        dt_frame_scrollbar = ctk.CTkScrollbar(frame)
        dt_frame_scrollbar.pack(side='right', fill='y')
        # Create the Treeview with the configured style
        dt_frame_treeview = ttk.Treeview(frame, columns=dt_frame_columns, selectmode='extended',
                                         show="headings", style="Treeview",
                                         yscrollcommand=dt_frame_scrollbar.set)

        dt_frame_treeview.tag_configure('oddrow', background='white')
        dt_frame_treeview.tag_configure('evenrow', background='lightblue')

        dt_frame_scrollbar.configure(command=dt_frame_treeview.yview)

        # Add column headings and data
        for col in dt_frame_columns:
            dt_frame_treeview.heading(col, text=col)
            dt_frame_treeview.column(col, width=int(master.winfo_width()*4/6*1/9), anchor='center')

        for row in dt_frame_data[1:]:
            row[6] = str(row[6])[:10]
            row[5] = str(row[5])[:10]

        count = 0
        for row in dt_frame_data[1:]:
            if count % 2:
                dt_frame_treeview.insert("", tk.END, values=row, tags="evenrow")
            else:
                dt_frame_treeview.insert("", tk.END, values=row, tags="oddrow")
            count += 1

        dt_frame_treeview.pack(fill=tk.BOTH, expand=True)

    frame.db1 = db1
    frame.db2 = db2
    frame.db3 = db3
    frame.dt_frame_treeview = dt_frame_treeview

    return frame


def create_tkinter_app():
    file_control(file, sheet1, trs_columns, sheet2, stk_columns, sheet3, hst_columns)

    file_data = data(file, sheet1)
    if len(file_data) > 500:
        file_output(file, sheet1, sheet2, sheet3)
        stock = data(file, sheet2)
        history = data(file, sheet3)
        os.remove(file)
        file_control(file, sheet1, trs_columns, sheet2, stk_columns, sheet3, hst_columns, stock, history)

    app = ctk.CTk()

    # Set the window title
    app._set_appearance_mode("dark")
    app.title("Stock Management System")
    w, h = app.winfo_screenwidth(), app.winfo_screenheight()
    app.geometry(f"{w}x{h}")

    # Set the grid columns and rows
    app.grid_columnconfigure(0, weight=1)
    app.grid_columnconfigure(1, weight=4)
    app.grid_rowconfigure(0, weight=1)
    app.grid_rowconfigure(1, weight=15)

    # Create the 3 frames

    de_frame = create_data_entry_frame(app)
    dt_frame = create_data_tables_frame(app)
    app.dt_frame = dt_frame
    button_frame = create_button_frame(app)

    # Store dt_frame as an attribute of app

    app.de_frame = de_frame
    app.button_frame = button_frame

    app.file_control = file_control
    app.data = data
    app.next_operation_code = next_operation_code
    app.next_product_code = next_product_code
    app.write_new_row = write_new_row
    app.overwrite = overwrite
    app.totals_from_zero = totals_from_zero
    app.new_to_totals = new_to_totals
    app.file_output = file_output
    app.delete_row = delete_row

    # Place the frames in the grid
    button_frame.grid(row=0, column=0, padx=10, pady=10, sticky="wens", rowspan=2)
    de_frame.grid(row=0, column=1, padx=10, pady=10, sticky="nswe")
    dt_frame.grid(row=1, column=1, padx=10, pady=10, sticky="nswe")

    # Start the main loop
    app.mainloop()


create_tkinter_app()
